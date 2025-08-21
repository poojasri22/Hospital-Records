from flask import Flask, render_template, request, redirect, url_for, session
import sqlite3
import io
import csv
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from flask import Flask, render_template, request, redirect, url_for, session, Response


app = Flask(__name__)
DATABASE = 'hospital.db'
app.secret_key = "supersecretkey"


# ===================== INITIALIZE DB ======================
def init_db():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    # Doctors Table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS doctors (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            specialization TEXT,
            phone TEXT
        )
    ''')

    # Patients Table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS patients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            age INTEGER,
            illness TEXT,
            doctor_id INTEGER,
            FOREIGN KEY (doctor_id) REFERENCES doctors(id)
        )
    ''')

    # Appointments Table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS appointments (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            patient_id INTEGER,
            doctor_id INTEGER,
            appointment_date TEXT,
            notes TEXT,
            FOREIGN KEY (patient_id) REFERENCES patients(id),
            FOREIGN KEY (doctor_id) REFERENCES doctors(id)
        )
    ''')

    conn.commit()
    conn.close()


# ===================== LOGIN / LOGOUT ======================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        if request.form['username'] == "admin" and request.form['password'] == "admin123":
            session['user'] = "admin"
            return redirect(url_for('dashboard'))
        else:
            return "Invalid credentials"
    return render_template('login.html')


@app.route('/logout')
def logout():
    session.pop('user', None)
    return redirect(url_for('login'))


# ===================== DASHBOARD ======================
@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        return redirect(url_for('login'))

    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("SELECT COUNT(*) FROM patients")
    total_patients = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM doctors")
    total_doctors = cursor.fetchone()[0]
    cursor.execute("SELECT COUNT(*) FROM appointments")
    total_appointments = cursor.fetchone()[0]
    conn.close()

    return render_template("dashboard.html",
                           total_patients=total_patients,
                           total_doctors=total_doctors,
                           total_appointments=total_appointments)


# ===================== PATIENTS CRUD ======================
@app.route('/patients')
def patients():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT patients.id, patients.name, patients.age, patients.illness, doctors.name
        FROM patients
        LEFT JOIN doctors ON patients.doctor_id = doctors.id
    """)
    patients = cursor.fetchall()

    cursor.execute("SELECT id, name FROM doctors")
    doctors = cursor.fetchall()

    conn.close()
    return render_template('patients.html', patients=patients, doctors=doctors)


@app.route('/add_patient', methods=['POST'])
def add_patient():
    name = request.form['name']
    age = request.form['age']
    illness = request.form['illness']
    doctor_id = request.form['doctor_id']
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO patients (name, age, illness, doctor_id) VALUES (?, ?, ?, ?)",
                   (name, age, illness, doctor_id))
    conn.commit()
    conn.close()
    return redirect(url_for('patients'))


@app.route('/delete_patient/<int:patient_id>')
def delete_patient(patient_id):
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM patients WHERE id=?", (patient_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('patients'))


@app.route('/export/csv')
def export_csv():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, age, illness FROM patients")
    data = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Name', 'Age', 'Illness'])
    writer.writerows(data)

    response = Response(output.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=patients.csv"
    return response


@app.route('/export/excel')
def export_excel():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, age, illness FROM patients")
    data = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Patients"
    headers = ['ID', 'Name', 'Age', 'Illness']
    ws.append(headers)

    for row in data:
        ws.append(row)

    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 15

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = Response(output.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = "attachment; filename=patients.xlsx"
    return response


@app.route('/export/pdf')
def export_pdf():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, age, illness FROM patients")
    data = cursor.fetchall()
    conn.close()

    output = io.BytesIO()
    p = canvas.Canvas(output, pagesize=letter)
    p.drawString(100, 750, "Patient Records")

    y = 700
    for row in data:
        p.drawString(100, y, f"ID: {row[0]}, Name: {row[1]}, Age: {row[2]}, Illness: {row[3]}")
        y -= 20
        if y < 50:
            p.showPage()
            y = 750

    p.save()
    output.seek(0)

    response = Response(output.read(), mimetype="application/pdf")
    response.headers["Content-Disposition"] = "attachment; filename=patients.pdf"
    return response

# ===================== EXPORT DOCTORS ======================
@app.route('/export/doctors/csv')
def export_doctors_csv():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, specialization, phone FROM doctors")
    data = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Name', 'Specialization', 'Phone'])
    writer.writerows(data)

    response = Response(output.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=doctors.csv"
    return response


@app.route('/export/doctors/excel')
def export_doctors_excel():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, specialization, phone FROM doctors")
    data = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Doctors"
    headers = ['ID', 'Name', 'Specialization', 'Phone']
    ws.append(headers)
    for row in data:
        ws.append(row)

    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = Response(output.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = "attachment; filename=doctors.xlsx"
    return response


@app.route('/export/doctors/pdf')
def export_doctors_pdf():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("SELECT id, name, specialization, phone FROM doctors")
    data = cursor.fetchall()
    conn.close()

    output = io.BytesIO()
    p = canvas.Canvas(output, pagesize=letter)
    p.drawString(100, 750, "Doctor Records")

    y = 700
    for row in data:
        p.drawString(100, y, f"ID: {row[0]}, Name: {row[1]}, Specialization: {row[2]}, Phone: {row[3]}")
        y -= 20
        if y < 50:
            p.showPage()
            y = 750

    p.save()
    output.seek(0)

    response = Response(output.read(), mimetype="application/pdf")
    response.headers["Content-Disposition"] = "attachment; filename=doctors.pdf"
    return response



# ===================== DOCTORS CRUD ======================
@app.route('/doctors')
def doctors():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM doctors")
    doctors = cursor.fetchall()
    conn.close()
    return render_template('doctors.html', doctors=doctors)


@app.route('/add_doctor', methods=['GET', 'POST'])
def add_doctor():
    if request.method == 'POST':
        name = request.form['name']
        specialization = request.form['specialization']
        phone = request.form['phone']
        conn = sqlite3.connect(DATABASE)
        cursor = conn.cursor()
        cursor.execute("INSERT INTO doctors (name, specialization, phone) VALUES (?, ?, ?)",
                       (name, specialization, phone))
        conn.commit()
        conn.close()
        return redirect(url_for('doctors'))

    # If GET request → show form
    return render_template('add_doctor.html')


@app.route('/delete_doctor/<int:doctor_id>')
def delete_doctor(doctor_id):
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM doctors WHERE id=?", (doctor_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('doctors'))


# ===================== APPOINTMENTS CRUD ======================
@app.route('/appointments')
def appointments():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT appointments.id, patients.name, doctors.name, appointments.appointment_date, appointments.notes
        FROM appointments
        JOIN patients ON appointments.patient_id = patients.id
        JOIN doctors ON appointments.doctor_id = doctors.id
    """)
    appointments = cursor.fetchall()

    cursor.execute("SELECT id, name FROM patients")
    patients = cursor.fetchall()
    cursor.execute("SELECT id, name FROM doctors")
    doctors = cursor.fetchall()

    conn.close()
    return render_template('appointments.html', appointments=appointments, patients=patients, doctors=doctors)


@app.route('/add_appointment', methods=['POST'])
def add_appointment():
    patient_id = request.form['patient_id']
    doctor_id = request.form['doctor_id']
    date = request.form['appointment_date']
    notes = request.form['notes']
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("INSERT INTO appointments (patient_id, doctor_id, appointment_date, notes) VALUES (?, ?, ?, ?)",
                   (patient_id, doctor_id, date, notes))
    conn.commit()
    conn.close()
    return redirect(url_for('appointments'))


@app.route('/delete_appointment/<int:appointment_id>')
def delete_appointment(appointment_id):
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("DELETE FROM appointments WHERE id=?", (appointment_id,))
    conn.commit()
    conn.close()
    return redirect(url_for('appointments'))

# ===================== EXPORT APPOINTMENTS ======================
@app.route('/export/appointments/csv')
def export_appointments_csv():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT appointments.id, patients.name, doctors.name, appointments.appointment_date, appointments.notes
        FROM appointments
        JOIN patients ON appointments.patient_id = patients.id
        JOIN doctors ON appointments.doctor_id = doctors.id
    """)
    data = cursor.fetchall()
    conn.close()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['ID', 'Patient', 'Doctor', 'Date', 'Notes'])
    writer.writerows(data)

    response = Response(output.getvalue(), mimetype="text/csv")
    response.headers["Content-Disposition"] = "attachment; filename=appointments.csv"
    return response


@app.route('/export/appointments/excel')
def export_appointments_excel():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT appointments.id, patients.name, doctors.name, appointments.appointment_date, appointments.notes
        FROM appointments
        JOIN patients ON appointments.patient_id = patients.id
        JOIN doctors ON appointments.doctor_id = doctors.id
    """)
    data = cursor.fetchall()
    conn.close()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Appointments"
    headers = ['ID', 'Patient', 'Doctor', 'Date', 'Notes']
    ws.append(headers)
    for row in data:
        ws.append(row)

    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 20

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = Response(output.read(), mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = "attachment; filename=appointments.xlsx"
    return response


@app.route('/export/appointments/pdf')
def export_appointments_pdf():
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()
    cursor.execute("""
        SELECT appointments.id, patients.name, doctors.name, appointments.appointment_date, appointments.notes
        FROM appointments
        JOIN patients ON appointments.patient_id = patients.id
        JOIN doctors ON appointments.doctor_id = doctors.id
    """)
    data = cursor.fetchall()
    conn.close()

    output = io.BytesIO()
    p = canvas.Canvas(output, pagesize=letter)
    p.drawString(100, 750, "Appointment Records")

    y = 700
    for row in data:
        p.drawString(100, y, f"ID: {row[0]}, Patient: {row[1]}, Doctor: {row[2]}, Date: {row[3]}, Notes: {row[4]}")
        y -= 20
        if y < 50:
            p.showPage()
            y = 750

    p.save()
    output.seek(0)

    response = Response(output.read(), mimetype="application/pdf")
    response.headers["Content-Disposition"] = "attachment; filename=appointments.pdf"
    return response


# ===================== SEARCH ======================
def search_data(query, start_date, end_date):
    conn = sqlite3.connect(DATABASE)
    cursor = conn.cursor()

    # Patients
    cursor.execute("""
        SELECT patients.id, patients.name, patients.age, patients.illness, doctors.name
        FROM patients
        LEFT JOIN doctors ON patients.doctor_id = doctors.id
        WHERE patients.name LIKE ? OR doctors.name LIKE ? OR patients.illness LIKE ?
    """, (f"%{query}%", f"%{query}%", f"%{query}%"))
    patient_results = cursor.fetchall()

    # Appointments
    sql = """
        SELECT appointments.id, patients.name, doctors.name, appointment_date, notes
        FROM appointments
        JOIN patients ON appointments.patient_id = patients.id
        JOIN doctors ON appointments.doctor_id = doctors.id
        WHERE (patients.name LIKE ? OR doctors.name LIKE ? OR appointment_date LIKE ? OR notes LIKE ?)
    """
    params = [f"%{query}%", f"%{query}%", f"%{query}%", f"%{query}%"]

    if start_date and end_date:
        sql += " AND appointment_date BETWEEN ? AND ?"
        params.extend([start_date, end_date])

    cursor.execute(sql, params)
    appointment_results = cursor.fetchall()

    conn.close()
    return patient_results, appointment_results


@app.route('/search')
def search():
    query = request.args.get('q')
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    patient_results, appointment_results = search_data(query, start_date, end_date)
    return render_template('search.html',
                           query=query,
                           start_date=start_date,
                           end_date=end_date,
                           patient_results=patient_results,
                           appointment_results=appointment_results)
# ===================== HOME ======================
@app.route('/')
def home():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    else:
        return redirect(url_for('login'))



# ===================== MAIN ======================
if __name__ == '__main__':
    init_db()
    app.run(debug=True)
